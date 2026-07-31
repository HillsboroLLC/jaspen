import io
import importlib
import zipfile

import pytest
from openpyxl import load_workbook


@pytest.fixture
def export_routes(app):
    return importlib.import_module("app.routes.export")


def _sample_wbs():
    return {
        "name": "Launch Readiness Plan",
        "start_date": "2026-08-03",
        "phases": [
            {"name": "Discovery", "task_ids": ["task-discovery"]},
            {"name": "Delivery", "task_ids": ["task-build", "task-launch"]},
        ],
        "tasks": [
            {
                "id": "task-discovery",
                "title": "Confirm launch requirements",
                "phase": "Discovery",
                "status": "done",
                "priority": "high",
                "owner": "Avery",
                "suggested_role": "Product Manager",
                "start_date": "2026-08-03",
                "due_date": "2026-08-05",
                "estimated_days": 3,
                "depends_on": [],
                "description": "Confirm scope, constraints, and acceptance criteria.",
                "risk_area": "execution_readiness",
                "rationale": "The delivery team needs one approved scope.",
                "order": 1,
                "external_refs": {},
            },
            {
                "id": "task-build",
                "title": "Build the launch package",
                "phase": "Delivery",
                "status": "in_progress",
                "priority": "medium",
                "owner": "Jordan",
                "suggested_role": "Delivery Lead",
                "start_date": "2026-08-06",
                "due_date": "2026-08-12",
                "timeline_days": 7,
                "depends_on": ["task-discovery"],
                "description": "Create the final assets and operating checklist.",
                "function": "Operations",
                "activity_type": "delivery",
                "order": 2,
                "external_refs": {"jira_issue_key": "JAS-42"},
            },
            {
                "id": "task-launch",
                "title": "=NOT_A_FORMULA()",
                "phase": "Delivery",
                "status": "todo",
                "priority": "low",
                "owner": "",
                "suggested_role": "Program Manager",
                "start_date": "2026-08-13",
                "due_date": "2026-08-14",
                "estimated_days": 2,
                "depends_on": ["task-build"],
                "acceptance": "Launch approval recorded.",
                "order": 3,
                "external_refs": {},
            },
        ],
    }


def _sample_scorecard():
    return {
        "analysis_id": "idea-1",
        "project_name": "Launch Program",
        "jaspen_score": 72,
        "score_category": "Good",
        "component_scores": {
            "strategic_alignment": 82,
            "financial_viability": 68,
            "execution_readiness": 74,
            "risk_profile": 63,
        },
        "financial_impact": {"estimated_value": "$1.2M", "payback_period": "18 months"},
        "risks": ["Adoption may lag without an owner."],
        "recommendations": ["Confirm ownership before launch."],
        "updated_at": "2026-07-31T20:00:00Z",
        "scenario_variants": [],
        "dimensions": {},
        "executive_summary": "The launch is strategically sound but needs a named owner.",
        "rubric": None,
    }


def _load(payload):
    return load_workbook(io.BytesIO(payload), data_only=False)


def test_wbs_xlsx_has_overview_and_filterable_task_table(export_routes):
    payload = export_routes._wbs_xlsx_bytes(
        _sample_wbs(),
        project_name="Launch Program",
        workspace_name="Strategy Team",
    )

    workbook = _load(payload)
    assert workbook.sheetnames == ["Overview", "Tasks"]

    overview = workbook["Overview"]
    assert overview["A1"].value == "Launch Program"
    assert overview["B4"].value == "Launch Program"
    assert overview["B5"].value == "Strategy Team"
    assert overview.freeze_panes is None
    assert overview["B10"].value == "=COUNTA('Tasks'!$C$2:$C$4)"
    assert overview["B15"].value.startswith("=IFERROR")
    assert overview["B20"].value == "Discovery"
    assert overview["B21"].value == "Delivery"
    assert "'Tasks'!$B$2:$B$4" in overview["C20"].value
    assert overview["F20"].value == "2026-08-03 to 2026-08-05"
    assert overview["F21"].value == "2026-08-06 to 2026-08-14"
    assert "ExecutionTasks[" not in "".join(
        str(overview.cell(row=row, column=column).value or "")
        for row in range(1, overview.max_row + 1)
        for column in range(1, overview.max_column + 1)
    )

    tasks = workbook["Tasks"]
    assert tasks.freeze_panes is None
    assert list(tasks.tables) == ["ExecutionTasks"]
    assert tasks.tables["ExecutionTasks"].ref == "A1:T4"
    assert tasks.auto_filter.ref is None
    assert [tasks.cell(row=1, column=column).value for column in range(1, 21)] == [
        "Phase #", "Phase", "Task", "Status", "Priority", "Owner", "Suggested Role",
        "Start Date", "Due Date", "Duration (days)", "Dependencies",
        "Description / Acceptance Criteria", "Risk Area", "Rationale", "Jira Key",
        "Function", "Activity Type", "Task ID", "Dependency IDs", "Task Order",
    ]

    assert tasks["A2"].value == 1
    assert tasks["B2"].value == "Discovery"
    assert tasks["D2"].value == "Done"
    assert tasks["O2"].value in (None, "")
    assert tasks["A3"].value == 2
    assert tasks["K3"].value == "Confirm launch requirements"
    assert tasks["O3"].value == "JAS-42"
    assert tasks["C4"].value == "=NOT_A_FORMULA()"
    assert tasks["C4"].data_type == "s"
    assert tasks["C4"].quotePrefix is True
    assert tasks["L4"].value == "Launch approval recorded."
    assert tasks["H2"].number_format == "yyyy-mm-dd"
    assert tasks["I2"].number_format == "yyyy-mm-dd"
    assert len(tasks.data_validations.dataValidation) == 2
    assert sum(len(rules) for rules in tasks.conditional_formatting._cf_rules.values()) == 7

    # A table already owns its filter. A second worksheet-level autoFilter over
    # the same cells causes desktop Excel to repair the workbook and remove the
    # table, which breaks all formulas that reference it.
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        task_sheet_xml = archive.read("xl/worksheets/sheet2.xml")
        table_xml = archive.read("xl/tables/table1.xml")
    assert b"<autoFilter" not in task_sheet_xml
    assert b'<autoFilter ref="A1:T4"' in table_xml


def test_wbs_xlsx_returns_none_without_tasks(export_routes):
    assert export_routes._wbs_xlsx_bytes({"tasks": []}) is None


def test_free_user_can_export_excel_and_csv_for_selected_scorecard(
    client,
    auth_headers,
    monkeypatch,
    export_routes,
):
    selected_wbs = _sample_wbs()
    other_wbs = {
        "name": "Other Plan",
        "phases": [{"name": "Other", "task_ids": ["other-task"]}],
        "tasks": [{"id": "other-task", "title": "Wrong task", "phase": "Other"}],
    }
    monkeypatch.setattr(
        export_routes,
        "load_user_sessions",
        lambda _user_id: {"thread-1": {"session_id": "thread-1", "name": "Launch Program"}},
    )
    monkeypatch.setattr(
        export_routes,
        "_load_scenarios",
        lambda _user_id: {
            "thread-1": {
                "project_wbs": other_wbs,
                "wbs_by_scorecard": {"idea-1": selected_wbs},
            }
        },
    )

    excel_response = client.get(
        "/api/v1/export/threads/thread-1/wbs/xlsx?scorecard_id=idea-1",
        headers=auth_headers,
    )
    assert excel_response.status_code == 200
    assert excel_response.mimetype == export_routes.XLSX_MIMETYPE
    assert "launch-program-execution-plan.xlsx" in excel_response.headers["Content-Disposition"].lower()
    workbook = _load(excel_response.data)
    task_titles = [workbook["Tasks"].cell(row=row, column=3).value for row in range(2, 5)]
    assert "Confirm launch requirements" in task_titles
    assert "Wrong task" not in task_titles

    csv_response = client.get(
        "/api/v1/export/threads/thread-1/wbs/csv?scorecard_id=idea-1",
        headers=auth_headers,
    )
    assert csv_response.status_code == 200
    assert "Confirm launch requirements" in csv_response.get_data(as_text=True)
    assert "Wrong task" not in csv_response.get_data(as_text=True)


def test_free_user_can_download_scorecard_pdf_and_editable_powerpoint(
    client,
    auth_headers,
    monkeypatch,
    export_routes,
):
    monkeypatch.setattr(
        export_routes,
        "load_user_sessions",
        lambda _user_id: {"thread-1": {"session_id": "thread-1", "name": "Launch Program"}},
    )
    monkeypatch.setattr(
        export_routes,
        "_scorecard_record_for_export",
        lambda _session, _thread_id, scorecard_id=None, user_id=None: (_sample_scorecard(), None),
    )

    pdf_response = client.get(
        "/api/v1/export/threads/thread-1/scorecard/pdf?scorecard_id=idea-1",
        headers=auth_headers,
    )
    assert pdf_response.status_code == 200
    assert pdf_response.mimetype == export_routes.PDF_MIMETYPE
    assert pdf_response.data.startswith(b"%PDF-")
    assert "launch-program-scorecard.pdf" in pdf_response.headers["Content-Disposition"].lower()

    pptx_response = client.get(
        "/api/v1/export/threads/thread-1/scorecard/pptx?scorecard_id=idea-1",
        headers=auth_headers,
    )
    assert pptx_response.status_code == 200
    assert pptx_response.mimetype == export_routes.PPTX_MIMETYPE
    assert "launch-program-scorecard.pptx" in pptx_response.headers["Content-Disposition"].lower()

    # Every exported scorecard element is a native PowerPoint shape/text box,
    # not a flattened screenshot, so customers can edit the deck.
    with zipfile.ZipFile(io.BytesIO(pptx_response.data)) as archive:
        slide_names = sorted(name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        slide_xml = b"".join(archive.read(name) for name in slide_names)
        media_names = [name for name in archive.namelist() if name.startswith("ppt/media/")]
    assert len(slide_names) >= 2
    assert b"Launch Program" in slide_xml
    assert b"Score breakdown" in slide_xml
    assert media_names == []


def test_scorecard_exports_use_saved_overrides_and_one_authoritative_rubric(export_routes):
    session = {
        "session_id": "thread-1",
        "name": "Old thread name",
        "result": {
            "analysis_id": "idea-1",
            "project_name": "Original title",
            "jaspen_score": 75,
            "score_category": "Good",
            # Legacy fields can remain on upgraded records. They must not create
            # an extra score-breakdown slide when modern dimensions exist.
            "component_scores": {
                "financial_health": 0,
                "operational_efficiency": 0,
                "market_position": 0,
                "execution_readiness": 0,
            },
            "dimensions": {
                "deployment": {"label": "Deployment readiness", "score": 75},
                "relationship": {"label": "Existing relationship", "score": 75},
                "customer": {"label": "Customer readiness", "score": 75},
                "market": {"label": "Market economics", "score": 75},
                "repeatability": {"label": "Repeatability", "score": 75},
            },
            "rubric": {
                "criteria": [
                    {"key": "deployment", "label": "Deployment readiness"},
                    {"key": "relationship", "label": "Existing relationship"},
                    {"key": "customer", "label": "Customer readiness"},
                    {"key": "market", "label": "Market economics"},
                    {"key": "repeatability", "label": "Repeatability"},
                ]
            },
            "financial_impact": {
                "ebitda_at_risk": None,
                "numeric": {"ebitda_at_risk": None},
            },
            "top_risks": ["Original risk"],
            "recommendations": [],
            "executive_summary": "Original summary",
            "display_overrides": {
                "title": "Saved customer-facing title",
                "executive_summary": "Saved customer-facing summary",
                "top_risks": ["Saved customer-facing risk"],
                "custom_blocks": [
                    {
                        "id": "blk_mitigation",
                        "type": "text",
                        "heading": "Mitigation",
                        "body": "Partner with the Zone J team before launch.",
                    }
                ],
            },
        },
    }

    scorecard, error = export_routes._scorecard_record_for_export(
        session, "thread-1", scorecard_id="idea-1"
    )
    assert error is None
    assert scorecard["project_name"] == "Saved customer-facing title"
    assert scorecard["executive_summary"] == "Saved customer-facing summary"
    assert scorecard["risks"] == ["Saved customer-facing risk"]
    assert scorecard["custom_blocks"][0]["heading"] == "Mitigation"

    pptx = export_routes._pptx_bytes(scorecard)
    with zipfile.ZipFile(io.BytesIO(pptx)) as archive:
        slide_names = sorted(
            name for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        slide_xml = b"".join(archive.read(name) for name in slide_names)

    assert len(slide_names) == 3
    assert b"Mitigation" in slide_xml
    assert b"Partner with the Zone J team before launch." in slide_xml
    assert b"Deployment readiness" in slide_xml
    assert b"Financial Health" not in slide_xml
    assert b"Numeric" not in slide_xml
